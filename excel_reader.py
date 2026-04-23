"""
excel_reader.py
────────────────────────────────────────────────────────────────
Backend reader for the Data Fusion Ontology ingestion pipeline.
Called by uploader.html (via Flask) OR used directly in Python.

Two modes:
  1. Full Excel upload  → parse all 3 sheets automatically
  2. Per-collection     → accepts a dict for a single collection

Usage (standalone):
    from excel_reader import ExcelReader
    reader = ExcelReader("path/to/file.xlsx")
    reader.load_all()   # inserts everything into MongoDB

Usage (from Flask / uploader.html):
    python server.py    # launches the Flask API + serves uploader.html
"""

import re
import hashlib
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook


# ── Helpers ────────────────────────────────────────────────────────────────

def _clean(value):
    """Strip whitespace and return None for empty/None values."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _split_csv(value):
    """Turn a comma-separated string into a cleaned list."""
    if not value:
        return []
    return [item.strip() for item in str(value).split(",") if item.strip()]


def _parse_date(value):
    """Return ISO date string from datetime object or raw string."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    cleaned = _clean(value)
    return cleaned  # leave as-is if already a string


# ── Sheet Parsers ───────────────────────────────────────────────────────────

def parse_papers_sheet(ws) -> list[dict]:
    """
    Parse the DOI sheet → list of paper documents.
    Expected columns (0-indexed):
      0  DOI            1  Title          2  Author
      3  Publication    4  Date           5  URL
      6  Keywords       7  Abstract       8  Publisher
      9  Field          10 IsDataFusion   11 ClassificationReason
    """
    rows = list(ws.iter_rows(values_only=True))
    docs = []

    for row in rows[2:]:           # skip header row + blank row
        doi = _clean(row[0])
        if not doi:
            continue               # skip empty rows

        doc = {
            "_id":                   doi,
            "title":                 _clean(row[1]),
            "authors":               _split_csv(row[2]),
            "publication_title":     _clean(row[3]),
            "publication_date":      _parse_date(row[4]),
            "url":                   _clean(row[5]),
            "keywords":              _split_csv(row[6]),
            "abstract":              _clean(row[7]),
            "publisher":             _clean(row[8]),
            "field_of_study":        _split_csv(row[9]),
            "is_data_fusion":        str(row[10]).strip().lower() == "yes" if row[10] else False,
            "classification_reason": _clean(row[11]),
        }
        docs.append(doc)

    return docs


def parse_fusion_methods_sheet(ws) -> list[dict]:
    """
    Parse the Fusion Method sheet → list of fusion method documents.
    Expected columns (0-indexed):
      0  Method Name    1  Method Key (UUID)   2  DOI
      3  Description    4  U1                  5  U3
    Note: U2 lives in the Data sheet (it's a dataset-level property).
    """
    rows = list(ws.iter_rows(values_only=True))
    docs = []

    for row in rows[2:]:
        method_name = _clean(row[0])
        if not method_name:
            continue

        doc = {
            "_id":         _clean(row[1]),       # UUID as _id
            "method_name": method_name,
            "paper_doi":   _clean(row[2]),        # FK → papers._id
            "description": _clean(row[3]),
            "uncertainty": {
                "U1_conception": _clean(row[4]),  # Ambiguity in world abstraction
                "U3_analysis":   _clean(row[5]),  # Algorithmic / processing errors
                # U2 is not here — it belongs to individual datasets
            },
            "dataset_ids": [],                    # back-filled during dataset import
        }
        docs.append(doc)

    return docs


def parse_datasets_sheet(ws) -> list[dict]:
    """
    Parse the Data sheet → list of dataset documents.
    Expected columns (0-indexed):
      0  DOI            1  Data Name       2  DatasetURL
      3  Method Key     4  Data Type       5  Collection Method
      6  U2             7  SpatialCoverage 8  TemporalCoverage
      9  Format         10 License         11 Provenance
    """
    rows = list(ws.iter_rows(values_only=True))
    docs = []

    for row in rows[2:]:
        data_name = _clean(row[1])
        if not data_name:
            continue

        # Stable _id: MD5 of "doi::data_name" — same row always maps to same id.
        # This makes re-uploading the same file (or overlapping files) safe:
        # MongoDB will upsert in place rather than inserting a duplicate row.
        paper_doi = _clean(row[0])
        stable_id = hashlib.md5(
            f"{paper_doi}::{data_name}".encode()
        ).hexdigest()

        doc = {
            "_id":                stable_id,
            "paper_doi":          paper_doi,         # FK → papers._id
            "data_name":          data_name,
            "dataset_url":        _clean(row[2]),
            "method_key":         _clean(row[3]),    # FK → fusion_methods._id
            "data_type":          _split_csv(row[4]),
            "collection_method":  _clean(row[5]),
            "uncertainty": {
                "U2_measurement": _clean(row[6]),    # Sensor errors / resolution limits
            },
            "spatial_coverage":   _clean(row[7]),
            "temporal_coverage":  _clean(row[8]),
            "format":             _clean(row[9]),
            "license":            _clean(row[10]),
            "provenance":         _clean(row[11]),
        }
        docs.append(doc)

    return docs


# ── Main Reader Class ───────────────────────────────────────────────────────

class ExcelReader:
    """
    Reads a Data Fusion Ontology Excel file and pushes data to MongoDB.

    Parameters
    ----------
    filepath : str | Path
        Path to the .xlsx file.
    mongo_uri : str
        MongoDB connection string (default: local).
    db_name : str
        Target database name.
    """

    SHEET_PARSERS = {
        "DOI":            ("papers",         parse_papers_sheet),
        "Fusion Method":  ("fusion_methods", parse_fusion_methods_sheet),
        "Data":           ("datasets",       parse_datasets_sheet),
    }

    def __init__(
        self,
        filepath: str | Path,
        mongo_uri: str = "mongodb://localhost:27017/",
        db_name: str = "data_fusion_ontology",
    ):
        self.filepath   = Path(filepath)
        self.mongo_uri  = mongo_uri
        self.db_name    = db_name
        self._wb        = None
        self._db        = None

    # ── File loading ──────────────────────────────────────────────

    def open(self):
        """Load the workbook (read-only for performance)."""
        self._wb = load_workbook(self.filepath, read_only=True, data_only=True)
        return self

    def close(self):
        if self._wb:
            self._wb.close()

    # ── MongoDB connection ────────────────────────────────────────

    def _get_db(self):
        """Lazy-connect to MongoDB."""
        if self._db is None:
            from pymongo import MongoClient
            client = MongoClient(self.mongo_uri)
            self._db = client[self.db_name]
        return self._db

    # ── Per-collection imports ────────────────────────────────────

    def load_papers(self) -> dict:
        """Parse DOI sheet and upsert into papers collection."""
        return self._load_sheet("DOI")

    def load_fusion_methods(self) -> dict:
        """Parse Fusion Method sheet and upsert into fusion_methods collection."""
        return self._load_sheet("Fusion Method")

    def load_datasets(self) -> dict:
        """
        Parse Data sheet, insert into datasets collection,
        and back-fill dataset_ids on each fusion_method document.

        Back-fill is scoped to THIS batch only (via batch_ids).
        Re-uploading the same file is safe: $addToSet prevents duplicates
        in dataset_ids, and the deterministic _id prevents duplicate rows.
        """
        result = self._load_sheet("Data")

        # Only back-fill the dataset IDs that came from this import batch
        db = self._get_db()
        batch_ids = result.pop("batch_ids", [])
        for doc in db["datasets"].find(
            {"_id": {"$in": batch_ids}}, {"_id": 1, "method_key": 1}
        ):
            if doc.get("method_key"):
                db["fusion_methods"].update_one(
                    {"_id": doc["method_key"]},
                    {"$addToSet": {"dataset_ids": doc["_id"]}},
                )

        return result

    def _load_sheet(self, sheet_name: str) -> dict:
        """Generic sheet loader — parse, upsert, return summary."""
        if not self._wb:
            self.open()

        if sheet_name not in self._wb.sheetnames:
            return {"error": f"Sheet '{sheet_name}' not found in workbook."}

        collection_name, parser_fn = self.SHEET_PARSERS[sheet_name]
        ws   = self._wb[sheet_name]
        docs = parser_fn(ws)

        if not docs:
            return {"collection": collection_name, "inserted": 0, "updated": 0}

        db         = self._get_db()
        collection = db[collection_name]
        inserted = updated = 0

        batch_ids = []
        for doc in docs:
            doc_id = doc.get("_id")
            if doc_id:
                res = collection.replace_one({"_id": doc_id}, doc, upsert=True)
                batch_ids.append(doc_id)
                if res.upserted_id:
                    inserted += 1
                else:
                    updated += 1
            else:
                # Fallback for any doc without an _id (should not occur after the fix)
                res = collection.insert_one(doc)
                batch_ids.append(res.inserted_id)
                inserted += 1

        return {
            "collection": collection_name,
            "inserted":   inserted,
            "updated":    updated,
            "total":      len(docs),
            "batch_ids":  batch_ids,   # used by load_datasets() back-fill
        }

    # ── Full import ───────────────────────────────────────────────

    def load_all(self) -> dict:
        """
        Import all three sheets in dependency order:
          papers → fusion_methods → datasets (with back-fill)
        """
        if not self._wb:
            self.open()

        results = {}
        results["papers"]         = self.load_papers()
        results["fusion_methods"] = self.load_fusion_methods()
        results["datasets"]       = self.load_datasets()

        self.close()
        return results


# ── CLI entry point ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys, json

    if len(sys.argv) < 2:
        print("Usage: python excel_reader.py <path_to_excel.xlsx>")
        sys.exit(1)

    reader = ExcelReader(sys.argv[1])
    results = reader.load_all()

    print("\n── Import Summary ───────────────────────")
    for collection, stats in results.items():
        print(f"  {collection}: {stats}")
    print("─────────────────────────────────────────\n")
